from django.shortcuts import render, redirect
from django.views.generic import ListView, TemplateView, View
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, F, Subquery, Prefetch
from datetime import timedelta
from datetime import datetime
from django.contrib import messages
from .utils.medicion import medir_rendimiento
from .utils.extracto import obtenerValorExtracto

from asociado.models import Asociado, ParametroAsociado, TarifaAsociado, ConveniosAsociado
from beneficiario.models import Mascota, Beneficiario
from historico.models import HistorialPagos, HistoricoAuxilio, HistoricoCredito
from ventas.models import DetalleVenta, HistoricoVenta, Producto
from parametro.models import FormaPago, MesTarifa, TipoAsociado
from .queries import obtenerNovedades, obtenerDescuentoNomina

#Funciones
from funciones.function import separarFecha, convertirFecha

#Libreria para generar el Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Clase base para los reportes en Excel
class BaseReporteExcel(View):

    nombre_hoja = "Reporte"
    titulo = "Reporte General"
    columnas = []
    ancho_columnas = []
    
    def get_queryset(self, request, *args, **kwargs):
        raise NotImplementedError("Debes implementar 'get_queryset' en la subclase.")

    def preparar_fila(self, obj):
        raise NotImplementedError("Debes implementar 'preparar_fila' en la subclase.")
    
    def generar_excel(self, queryset, titulo):
        wb = Workbook()
        ws = wb.active
        ws.title = self.nombre_hoja

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")
        bold_font_header = Font(bold=True, size=11, color="000000")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_header = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.columnas))
        ws.cell(row=1, column=1).value = titulo
        ws.cell(row=1, column=1).font = bold_font
        ws.cell(row=1, column=1).alignment = alignment_center
        ws.cell(row=1, column=1).fill = fill_header

        # Encabezados
        for idx, encabezado in enumerate(self.columnas, 1):
            celda = ws.cell(row=2, column=idx)
            celda.value = encabezado
            celda.font = bold_font_header
            celda.alignment = alignment_center

        # Ancho de columnas
        for idx, width in enumerate(self.ancho_columnas, start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

        # Llenar datos
        for i, obj in enumerate(queryset, start=3):
            fila = self.preparar_fila(obj)
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j).value = valor

        return wb

    # Se utiliza cuando es llamado la clase directamente con el get
    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset(request, *args, **kwargs)
        titulo = self.titulo
        if not queryset.exists():
            messages.error(request, "No se encontraron resultados. Verifica las fechas.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        wb = self.generar_excel(queryset, titulo)
        fecha = datetime.now().strftime("%d-%m-%Y")
        nombre_archivo = f"{self.nombre_hoja}_{fecha}.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
        wb.save(response)
        return response
    
    def exportar_excel(self, request, *args, **kwargs):
        queryset = self.get_queryset(request, *args, **kwargs)
        wb = self.generar_excel(queryset, self.titulo)
        fecha = datetime.now().strftime("%Y-%m-%d")
        nombre_archivo = f"{self.nombre_hoja}_{fecha}.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
        wb.save(response)
        return response
    
    # se utiliza cuando se llama a la clase desde un formulario POST, llamando a la funcion exportar_excel
    def post(self, request, *args, **kwargs):
        queryset = self.get_queryset(request, *args, **kwargs)
        if not queryset.exists():
            messages.error(request, "No se encontraron resultados. Verifica las fechas.")
            return render(request, 'mi_template.html', {
                'post': 'yes',
                # otras variables necesarias para que el template cargue bien
            })
        return self.exportar_excel(request, *args, **kwargs)

class InformacionReporte(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/informacion.html'
        return render(request, template_name)

class VerModificacionesFecha(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/modificacionesPorFecha.html'
        return render(request, template_name)
    
    def post(self, request, *args, **kwargs):
        fechaInicialForm = request.POST['fechaInicial']
        fechaFinalForm = request.POST['fechaFinal']
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario.
        try:
            fechaInicial = separarFecha(fechaInicialForm, 'inicial')
            fechaFinal = separarFecha(fechaFinalForm, 'final')
        except (ValueError, TypeError):
            messages.error(request, f"La fecha ingresada no es válida. Fecha Inicial {fechaInicialForm}, Fecha Final {fechaFinalForm}")
            return render(request, 'reporte/modificacionesPorFecha.html', {'post':'post', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm})
        
        # Obtenemos las queries de las novedades
        queries = obtenerNovedades(fechaInicial, fechaFinal)

        template_name = 'reporte/modificacionesPorFecha.html'
        context = {'queries':queries,
                'post':'yes',
                'fechaIncialF':fechaInicialForm,
                'fechaFinalF':fechaFinalForm,
                }
        return render(request, template_name, context)

class ReporteExcelFecha(TemplateView):
    def get(self, request, *args, **kwargs):
        fechaInicialForm = request.GET.get('fechaInicial')
        fechaFinalForm = request.GET.get('fechaFinal')
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario.
        try:
            fechaInicial = separarFecha(fechaInicialForm, 'inicial')
            fechaFinal = separarFecha(fechaFinalForm, 'final')
        except (ValueError, TypeError):
            messages.error(request, f"La fecha ingresada no es válida. Fecha Inicial {fechaInicialForm}, Fecha Final {fechaFinalForm}")
            return render(request, 'reporte/modificacionesPorFecha.html', {'post':'post', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm})

        # Funcion para formatear la fecha en el formato d-m-Y
        fecha_formateada1 = convertirFecha(fechaInicialForm)
        fecha_formateada2 = convertirFecha(fechaFinalForm)

        # Obtenemos las queries de las novedades
        queries = obtenerNovedades(fechaInicial, fechaFinal)

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")  # Fuente en negrita, tamaño 12 y color blanco
        bold_font2 = Font(bold=True, size=12, color="000000")  # Fuente en negrita, tamaño 12 y color negro
        alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación al centro
        fill = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")  # Relleno verde sólido

        wb = Workbook() #Creamos la instancia del Workbook
        ws = wb.active
        ws.title = 'Mascotas'
        titulo1 = f"Reporte de Novedades Mascota Funeraria desde {fecha_formateada1} - {fecha_formateada2}"
        ws['A1'] = titulo1    #Casilla en la que queremos poner la informacion
        ws.merge_cells('A1:I1')
        ws['A1'].font = bold_font
        ws['A1'].alignment = alignment_center
        ws['A1'].fill = fill

        ws['A2'] = 'Número registro'
        ws['B2'] = 'Número Documento Titular'
        ws['C2'] = 'Nombre Titular'
        ws['D2'] = 'Mascota'
        ws['E2'] = 'Tipo'
        ws['F2'] = 'Raza'
        ws['G2'] = 'Fecha Nacimiento'
        ws['H2'] = 'Novedad'
        ws['I2'] = 'Fecha Novedad'

        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,10):
            cell = ws.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 14
        
        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        for mascota in queries["queryMascota"]:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws.cell(row = cont, column = 1).value = i
            ws.cell(row = cont, column = 2).value = int(mascota.asociado.numDocumento)
            ws.cell(row = cont, column = 3).value = f'{mascota.asociado.nombre}' + ' ' + f'{mascota.asociado.apellido}'
            ws.cell(row = cont, column = 4).value = mascota.nombre
            ws.cell(row = cont, column = 5).value = mascota.tipo
            ws.cell(row = cont, column = 6).value = mascota.raza
            ws.cell(row = cont, column = 7).value = mascota.fechaNacimiento.strftime("%d/%m/%Y")
            if mascota.fechaCreacion_truncada == mascota.fechaModificacion_truncada:
                ws.cell(row = cont, column = 8).value = 'Ingreso'
                ws.cell(row = cont, column = 9).value = mascota.fechaIngreso.strftime("%d/%m/%Y")
            elif mascota.fechaCreacion_truncada != mascota.fechaModificacion_truncada:
                if mascota.estadoRegistro == True:
                    ws.cell(row = cont, column = 8).value = 'Modificación'
                    ws.cell(row = cont, column = 9).value = mascota.fechaModificacion.strftime("%d/%m/%Y")
                else:
                    ws.cell(row = cont, column = 8).value = 'Retiro'
                    ws.cell(row = cont, column = 9).value = mascota.fechaRetiro.strftime("%d/%m/%Y")
            i+=1
            cont+=1
        
        # se crea una nueva hoja
        wb.create_sheet('Beneficiarios')  
        # se selecciona la hoja creada
        ws2 = wb['Beneficiarios']
        titulo2 = f"Reporte de Novedades Beneficiarios Funeraria desde {fecha_formateada1} - {fecha_formateada2}"
        ws2['A1'] = titulo2    #Casilla en la que queremos poner la informacion
        ws2.merge_cells('A1:K1')
        ws2['A1'].font = bold_font
        ws2['A1'].alignment = alignment_center
        ws2['A1'].fill = fill

        ws2['A2'] = 'Número Registro'
        ws2['B2'] = 'Número Documento Titular'
        ws2['C2'] = 'Nombre Titular'
        ws2['D2'] = 'Tipo Documento Beneficiario'
        ws2['E2'] = 'Número Documento Beneficiario'
        ws2['F2'] = 'Nombre Beneficiario'
        ws2['G2'] = 'Parentesco'
        ws2['H2'] = 'Fecha Nacimiento'
        ws2['I2'] = 'Repatriación'
        ws2['J2'] = 'Novedad'
        ws2['K2'] = 'Fecha Novedad'

        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,13):
            cell = ws2.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws2.column_dimensions['A'].width = 11
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 40
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 13
        ws2.column_dimensions['F'].width = 40
        ws2.column_dimensions['G'].width = 16
        ws2.column_dimensions['H'].width = 12
        ws2.column_dimensions['I'].width = 15
        ws2.column_dimensions['J'].width = 13
        ws2.column_dimensions['K'].width = 12


        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        for beneficiario in queries["queryBeneficiario"]:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws2.cell(row = cont, column = 1).value = i                    
            ws2.cell(row = cont, column = 2).value = int(beneficiario.asociado.numDocumento)
            ws2.cell(row = cont, column = 3).value = f'{beneficiario.asociado.nombre}' + ' ' + f'{beneficiario.asociado.apellido}'
            ws2.cell(row = cont, column = 4).value = beneficiario.tipoDocumento
            ws2.cell(row = cont, column = 5).value = int(beneficiario.numDocumento)
            ws2.cell(row = cont, column = 6).value = f'{beneficiario.nombre}' + ' ' + f'{beneficiario.apellido}'
            ws2.cell(row = cont, column = 7).value = beneficiario.parentesco.nombre
            ws2.cell(row = cont, column = 8).value = beneficiario.fechaNacimiento.strftime("%d/%m/%Y")
            if beneficiario.repatriacion == True:
                ws2.cell(row = cont, column = 9).value = f'{beneficiario.paisRepatriacion.nombre}' + '-' + f'{beneficiario.ciudadRepatriacion}'
            else:
                ws2.cell(row = cont, column = 9).value = ''
            if beneficiario.fechaCreacion_truncada == beneficiario.fechaModificacion_truncada:
                ws2.cell(row = cont, column = 10).value = 'Ingreso'
            elif beneficiario.fechaCreacion_truncada != beneficiario.fechaModificacion_truncada:
                if beneficiario.estadoRegistro == True:
                    ws2.cell(row = cont, column = 10).value = 'Modificación'
                else:
                    ws2.cell(row = cont, column = 10).value = 'Retiro'
            if beneficiario.fechaRetiro != None:
                ws2.cell(row = cont, column = 11).value = beneficiario.fechaRetiro.strftime("%d/%m/%Y")
            elif beneficiario.fechaCreacion_truncada == beneficiario.fechaModificacion_truncada:
                ws2.cell(row = cont, column = 11).value = beneficiario.fechaIngreso.strftime("%d/%m/%Y")
            else:
                ws2.cell(row = cont, column = 11).value = beneficiario.fechaModificacion.strftime("%d/%m/%Y")
            i+=1
            cont+=1
        
        # se crea una nueva hoja
        wb.create_sheet('Asociados')
        # se selecciona la hoja creada
        ws3 = wb['Asociados']
        titulo2 = f"Reporte de Novedades Asociados Funeraria desde {fecha_formateada1} - {fecha_formateada2}"
        ws3['A1'] = titulo2    #Casilla en la que queremos poner la informacion
        ws3.merge_cells('A1:G1')

        ws3['A1'].font = bold_font
        ws3['A1'].alignment = alignment_center
        ws3['A1'].fill = fill

        ws3['A2'] = 'Número registro'
        ws3['B2'] = 'Nombre'
        ws3['C2'] = 'Tipo Documento'
        ws3['D2'] = 'Número Documento'
        ws3['E2'] = 'Fecha Nacimiento'
        ws3['F2'] = 'Novedad'
        ws3['G2'] = 'Fecha'

        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,8):
            cell = ws3.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws3.column_dimensions['A'].width = 11
        ws3.column_dimensions['B'].width = 30
        ws3.column_dimensions['C'].width = 14
        ws3.column_dimensions['D'].width = 13
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 29
        ws3.column_dimensions['G'].width = 12

        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        for asociado in queries["asociadoRetiro"]:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws3.cell(row = cont, column = 1).value = i                    
            ws3.cell(row = cont, column = 2).value = f'{asociado.nombre}' + ' ' + f'{asociado.apellido}'
            ws3.cell(row = cont, column = 3).value = asociado.tipoDocumento
            ws3.cell(row = cont, column = 4).value = int(asociado.numDocumento)
            ws3.cell(row = cont, column = 5).value = asociado.fechaNacimiento.strftime("%d/%m/%Y")
            ws3.cell(row = cont, column = 6).value = 'Retiro'
            ws3.cell(row = cont, column = 7).value = asociado.fechaRetiro.strftime("%d/%m/%Y")
            i+=1
            cont+=1
        
        for asociado in queries["asociadoIngreso"]:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws3.cell(row = cont, column = 1).value = i                    
            ws3.cell(row = cont, column = 2).value = f'{asociado.nombre}' + ' ' + f'{asociado.apellido}'
            ws3.cell(row = cont, column = 3).value = asociado.tipoDocumento
            ws3.cell(row = cont, column = 4).value = int(asociado.numDocumento)
            ws3.cell(row = cont, column = 5).value = asociado.fechaNacimiento.strftime("%d/%m/%Y")
            ws3.cell(row = cont, column = 6).value = 'Ingreso'
            ws3.cell(row = cont, column = 7).value = asociado.fechaIngreso.strftime("%d/%m/%Y")
            i+=1
            cont+=1

        for asociado in queries["queryRepatriacionTitular"]:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws3.cell(row = cont, column = 1).value = i                    
            ws3.cell(row = cont, column = 2).value = f'{asociado.asociado.nombre}' + ' ' + f'{asociado.asociado.apellido}'
            ws3.cell(row = cont, column = 3).value = asociado.asociado.tipoDocumento
            ws3.cell(row = cont, column = 4).value = int(asociado.asociado.numDocumento)
            ws3.cell(row = cont, column = 5).value = asociado.asociado.fechaNacimiento.strftime("%d/%m/%Y")
            if asociado.estadoRegistro == True:
                ws3.cell(row = cont, column = 6).value = 'Ingreso Repatriación ' + f'{asociado.paisRepatriacion}'
            else:
                ws3.cell(row = cont, column = 6).value = 'Retiro Repatriación ' + f'{asociado.paisRepatriacion}'
            ws3.cell(row = cont, column = 7).value = asociado.fechaRepatriacion.strftime("%d/%m/%Y")
            i+=1
            cont+=1
        

        nombre_archivo = f"Reporte Modificaciones_{fecha_formateada1}-{fecha_formateada2}.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class VerPagosFecha(TemplateView):
    template_name = 'reporte/pagosPorFecha.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)
    
    def post(self, request, *args, **kwargs):
        fechaInicialForm = request.POST['fechaInicial']
        fechaFinalForm = request.POST['fechaFinal']
        try:
            # Convertir fechas al formato datetime directamente
            fechaInicial = datetime.strptime(fechaInicialForm, "%Y-%m-%d")
            fechaFinal = datetime.strptime(fechaFinalForm, "%Y-%m-%d")
        except (ValueError, TypeError):
            messages.error(request, f"La fecha ingresada no es válida. Fecha Inicial {fechaInicialForm}, Fecha Final {fechaFinalForm}")
            return render(request, 'reporte/pagosPorFecha.html', {'post':'post', 'fechaInicialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm})

        queryHistorial = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal]
        ).select_related('asociado', 'mesPago', 'formaPago')

        totalPago = queryHistorial.aggregate(total=Sum('valorPago'))['total'] or 0

        context = {
            'query': queryHistorial,
            'post': 'yes',
            'fechaInicialF': fechaInicialForm,
            'fechaFinalF': fechaFinalForm,
            'totalPago': totalPago
        }

        return render(request, self.template_name, context)
    
class ReporteExcelPago(BaseReporteExcel):
    nombre_hoja = "Pagos"
    columnas = [
        'Número Registro', 'Fecha Pago','Número Documento', 'Nombre Completo', 'Mes Pago',
        'Valor Pago', 'Aporte', 'Bienestar Social', 'Mascota', 'Repatriación',
        'Seguro Vida', 'Adicionales', 'Coohoperativitos', 'Diferencia', 'Forma Pago'
    ]
    ancho_columnas = [11, 14, 14, 30, 16, 14, 12, 12, 12, 12, 12, 12, 16, 12, 12]

    def get_queryset(self, request, *args, **kwargs):
        fechaInicialForm = request.GET['fechaInicial']
        fechaFinalForm = request.GET['fechaFinal']
        try:
            fechaInicial = datetime.strptime(fechaInicialForm, "%Y-%m-%d")
            fechaFinal = datetime.strptime(fechaFinalForm, "%Y-%m-%d")
        except (ValueError, TypeError):
            messages.error(request, f"La fecha ingresada no es válida. Fecha Inicial {fechaInicialForm}, Fecha Final {fechaFinalForm}")
            return HistorialPagos.objects.none()

        self.titulo = f"Reporte de Pagos desde {fechaInicial.strftime('%d-%m-%Y')} hasta {fechaFinal.strftime('%d-%m-%Y')}"

        return HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal]
        ).select_related('asociado', 'mesPago', 'formaPago')

    def preparar_fila(self, obj):
        return [
            obj.id,
            obj.fechaPago.strftime("%d/%m/%Y"),
            int(obj.asociado.numDocumento),
            f'{obj.asociado.nombre} {obj.asociado.apellido}',
            obj.mesPago.concepto,
            obj.valorPago,
            obj.aportePago,
            obj.bSocialPago,
            obj.mascotaPago,
            obj.repatriacionPago,
            obj.seguroVidaPago,
            obj.adicionalesPago,
            obj.coohopAporte + obj.coohopBsocial,
            obj.diferencia,
            obj.formaPago.formaPago
        ]

class FormatoExtracto(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/reporteExtracto.html'
        mes = MesTarifa.objects.filter(pk__lt = 9000)
        return render(request, template_name, {'mes': mes})

    def post(self, request, *args, **kwargs):
        """
        Mostrar template inmediatamente sin procesar datos
        """
        template_name = 'reporte/generarExtracto.html'
        
        mesExtracto = request.POST.get('mesExtracto')
        saldos = 'saldos' in request.POST
        
        mes = MesTarifa.objects.get(pk=mesExtracto)
        
        # Renderizar template inmediatamente (sin datos)
        return render(request, template_name, {
            'mesExtracto': mesExtracto,
            'saldos': 'true' if saldos else 'false',
            'mes': mes.concepto
        })


class ObtenerExtractosAPI(View):
    """
    API endpoint para obtener extractos procesados
    """
    @medir_rendimiento('obtener_extractos_api')
    def post(self, request, *args, **kwargs):
        try:
            mesExtracto = request.POST.get('mesExtracto')
            saldos = request.POST.get('saldos') == 'true'
            
            # Obtener asociados
            objAsoc = (
                Asociado.objects
                .exclude(estadoAsociado = 'RETIRO')
                .exclude(parametroasociado__primerMes__id=9989)
                .select_related('mpioResidencia'))
            total = objAsoc.count()
            mes = MesTarifa.objects.get(pk=mesExtracto)
            
            extractos_lista = []
            errores = []
            
            for i, asociado in enumerate(objAsoc, 1):
                try:
                    # Generar extracto
                    context = obtenerValorExtracto(asociado.id, saldos, mes)
                    
                    extracto_serializable = {
                        "pkAsociado": asociado.id,
                        "nombre": f"{asociado.nombre} {asociado.apellido}",
                        "numDocumento": asociado.numDocumento,
                        "mpioResidencia": str(asociado.mpioResidencia),
                        "direccion": asociado.direccion,
                        "numCelular": asociado.numCelular,
                        "fechaCorte": context["fechaCorte"].strftime("%Y-%m-%d"),
                        "mes": context["mes"].concepto,
                        "saldo": context["saldo"],
                        "saldoDiferencia": context["saldoDiferencia"],
                        "pagoTotal": context["pagoTotal"],
                        "mensaje": context["mensaje"],
                        "conceptos_detallados": context["conceptos_detallados"],
                        "beneficiarios": [
                            {
                                "nombre": f"{b.nombre} {b.apellido}",
                                "parentesco": str(b.parentesco),
                                "paisRepatriacion": str(b.paisRepatriacion) if b.paisRepatriacion else "",
                            }
                            for b in context["objBeneficiario"]
                        ],
                        "mascotas": [
                            {"nombre": m.nombre, "tipo": m.tipo}
                            for m in context["objMascota"]
                        ],
                    }
                    
                    extractos_lista.append(extracto_serializable)
                    
                    # Enviar progreso periódicamente (cada 10 asociados)
                    if i % 10 == 0:
                        print(f"Progreso: {i}/{total} asociados procesados")
                    
                except Exception as e:
                    errores.append({
                        'asociado_id': asociado.id,
                        'nombre': f"{asociado.nombre} {asociado.apellido}",
                        'error': str(e)
                    })
                    continue
            
            return JsonResponse({
                'success': True,
                'extractos': extractos_lista,
                'total': len(extractos_lista),
                'errores': errores
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class VerDescuentosNomina(ListView):
    template_name = 'reporte/dctosNomina.html'
    
    def get(self, request, *args, **kwargs):
        empresas = TipoAsociado.objects.all()
        return render(request, self.template_name, {'empresas':empresas})
    
    def post(self, request, *args, **kwargs):
        empresas = TipoAsociado.objects.all()
        array = []
        arrayEmp = []
        granTotalGeneral = 0

        for empresa in empresas:
            if len(request.POST.getlist('select'+str(empresa.pk))) == 1:
                query = obtenerDescuentoNomina(empresa.pk)
                array.extend(query['query'])
                arrayEmp.append(empresa.pk)
                granTotalGeneral += query['granTotal']

        return render(request, self.template_name,{'array':array, 'post':'yes', 'empresas':empresas, 'arrayEmp':arrayEmp, 'granTotalGeneral':granTotalGeneral})

class ExcelDescuentosNomina(TemplateView):
    def get(self, request, *args, **kwargs):
        empresas_ids = [key.replace("select", "") for key in request.GET.keys() if key.startswith("select")]
        
        if not empresas_ids:
            return HttpResponse("No se seleccionaron empresas.", content_type="text/plain")

        array = []
        for empresa_id in empresas_ids:
            query = obtenerDescuentoNomina(empresa_id)
            array.extend(query['query'])

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")  # Fuente en negrita, tamaño 12 y color blanco
        bold_font2 = Font(bold=True, size=12, color="000000")  # Fuente en negrita, tamaño 12 y color negro
        alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación al centro
        fill = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")  # Relleno verde sólido

        wb = Workbook() #Creamos la instancia del Workbook
        ws = wb.active
        ws.title = 'Descuentos'
        titulo1 = f"Reporte descuentos nomina"
        ws['A1'] = titulo1    #Casilla en la que queremos poner la informacion
        ws.merge_cells('A1:S1')
        ws['A1'].font = bold_font
        ws['A1'].alignment = alignment_center
        ws['A1'].fill = fill

        ws['A2'] = 'Número registro'
        ws['B2'] = 'Codigo'
        ws['C2'] = 'Número Documento'
        ws['D2'] = 'Nombre Completo'
        ws['E2'] = 'Empresa'
        ws['F2'] = 'Valor'
        ws['G2'] = 'Aporte'
        ws['H2'] = 'Bienestar Social'
        ws['I2'] = 'Mascota'
        ws['J2'] = 'Repatriación'
        ws['K2'] = 'Seguro Vida'
        ws['L2'] = 'Adicionales'
        ws['M2'] = 'Coohoperativitos Aporte'
        ws['N2'] = 'Coohoperativitos B Social'
        ws['O2'] = 'Convenios'
        ws['P2'] = 'Convenio Gasolina'
        ws['Q2'] = 'Descuento Vinculación'
        ws['R2'] = 'Descuento Crédito'
        ws['S2'] = 'Descuento Crédito Home Elements'

        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,20):
            cell = ws.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['B'].width = 11
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 36
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 14
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 14
        ws.column_dimensions['O'].width = 14
        ws.column_dimensions['P'].width = 14
        ws.column_dimensions['Q'].width = 14
        ws.column_dimensions['R'].width = 14
        ws.column_dimensions['S'].width = 14

        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        for obj in array:
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws.cell(row = cont, column = 1).value = i
            ws.cell(row = cont, column = 2).value = obj.asociado.pk
            ws.cell(row = cont, column = 3).value = int(obj.asociado.numDocumento)
            ws.cell(row = cont, column = 4).value = f'{obj.asociado.nombre}' + ' ' + f'{obj.asociado.apellido}'
            ws.cell(row = cont, column = 5).value = obj.asociado.tAsociado.concepto
            ws.cell(row = cont, column = 6).value = obj.total_final
            ws.cell(row = cont, column = 7).value = obj.tarifaAsociado.cuotaAporte
            ws.cell(row = cont, column = 8).value = obj.tarifaAsociado.cuotaBSocial
            ws.cell(row = cont, column = 9).value = obj.tarifaAsociado.cuotaMascota
            ws.cell(row = cont, column = 10).value = (obj.tarifaAsociado.cuotaRepatriacionBeneficiarios or 0) + (obj.tarifaAsociado.cuotaRepatriacionTitular or 0)
            ws.cell(row = cont, column = 11).value = obj.tarifaAsociado.cuotaSeguroVida
            ws.cell(row = cont, column = 12).value = obj.tarifaAsociado.cuotaAdicionales
            ws.cell(row = cont, column = 13).value = obj.tarifaAsociado.cuotaCoohopAporte
            ws.cell(row = cont, column = 14).value = obj.tarifaAsociado.cuotaCoohopBsocial
            ws.cell(row = cont, column = 15).value = obj.tarifaAsociado.cuotaConvenio
            ws.cell(row = cont, column = 16).value = obj.cuota_convenio_gasolina
            ws.cell(row = cont, column = 17).value = obj.cuota_vinculacion if obj.cuota_vinculacion else 0
            ws.cell(row = cont, column = 18).value = obj.cuota_credito if obj.cuota_credito else 0
            ws.cell(row = cont, column = 19).value = obj.cuota_credito_home_elements if obj.cuota_credito_home_elements else 0
            i+=1
            cont+=1

        nombre_archivo = f"Reporte_Descuento_Nomina.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
class VerConciliacionBancaria(ListView):
    def get(self, request, *args, **kwargs):
        template_name = 'reporte/listarConciliacionBancaria.html'
        formaPago = FormaPago.objects.all().order_by('formaPago')
        return render(request, template_name, {'formaPago': formaPago})
    
    def post(self, request, *args, **kwargs):
        formaPago = FormaPago.objects.all().order_by('formaPago')
        fechaInicialForm = request.POST['fechaInicial']
        fechaFinalForm = request.POST['fechaFinal']
        
        try:
            fechaInicial = separarFecha(fechaInicialForm, 'inicial')
            fechaFinal = separarFecha(fechaFinalForm, 'final')
        except (ValueError, TypeError):
            messages.error(request, f"La fecha ingresada no es válida. Fecha Inicial {fechaInicialForm}, Fecha Final {fechaFinalForm}")
            formaPago = FormaPago.objects.all().order_by('formaPago')
            return render(request, 'reporte/listarConciliacionBancaria.html', {'post':'post', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm, 'formaPago':formaPago})
        
        banco = int(request.POST['banco'])
        # funcion que separa año, mes y dia de la fecha ingresada por el usuario.
        # consulta por rango de fecha inicial y final
        queryHistorial = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal],
            formaPago_id=banco
            ).values('asociado__id',
                'asociado__nombre',
                'asociado__apellido',
                'asociado__numDocumento',
                'formaPago_id__formaPago',
                'fechaPago',
            ).annotate(total_pagado=Sum('valorPago')).order_by('fechaPago')
        
        totalPago = queryHistorial.aggregate(total=Sum('total_pagado'))['total'] or 0
            
        template_name = 'reporte/listarConciliacionBancaria.html'
        return render(request, template_name, {'query':queryHistorial,'post':'post', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm, 'formaPago':formaPago, 'banco':banco, 'totalPago':totalPago})

class ExcelConciliacionBancaria(TemplateView):
    def get(self, request, *args, **kwargs):
        fechaInicialForm = request.GET['fechaInicial']
        fechaFinalForm = request.GET['fechaFinal']
        banco = int(request.GET['banco'])
        
        try:
            fechaInicial = separarFecha(fechaInicialForm, 'inicial')
            fechaFinal = separarFecha(fechaFinalForm, 'final')
        except (ValueError, TypeError):
            messages.error(request, f"La fecha ingresada no es válida. Fecha Inicial {fechaInicialForm}, Fecha Final {fechaFinalForm}")
            formaPago = FormaPago.objects.all().order_by('formaPago')
            return render(request, 'reporte/listarConciliacionBancaria.html', {'post':'post', 'fechaIncialF':fechaInicialForm, 'fechaFinalF':fechaFinalForm, 'formaPago':formaPago})
        
        # consulta por rango de fecha inicial y final
        queryHistorial = HistorialPagos.objects.filter(
            fechaPago__range=[fechaInicial, fechaFinal],
            formaPago_id=banco
            ).values('asociado__id',
                    'asociado__nombre',
                    'asociado__apellido',
                    'asociado__numDocumento',
                    'formaPago_id__formaPago',
                    'fechaPago',
            ).annotate(total_pagado=Sum('valorPago')).order_by('fechaPago')

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")  # Fuente en negrita, tamaño 12 y color blanco
        bold_font2 = Font(bold=True, size=12, color="000000")  # Fuente en negrita, tamaño 12 y color negro
        alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación al centro
        fill = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")  # Relleno verde sólido

        wb = Workbook() #Creamos la instancia del Workbook
        ws = wb.active
        ws.title = 'Conciliación Bancaria'
        titulo1 = f"Conciliación Bancaria"
        ws['A1'] = titulo1    #Casilla en la que queremos poner la informacion
        ws.merge_cells('A1:F1')
        ws['A1'].font = bold_font
        ws['A1'].alignment = alignment_center
        ws['A1'].fill = fill

        ws['A2'] = 'Número registro'
        ws['B2'] = 'Número Documento'
        ws['C2'] = 'Nombre Completo'
        ws['D2'] = 'Movimiento'
        ws['E2'] = 'Valor'
        ws['F2'] = 'Fecha Movimiento'

        bold_font2 = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1,7):
            cell = ws.cell(row=2, column=col)
            cell.font = bold_font2
            cell.alignment = center_alignment

        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 36
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14

        #Inicia el primer registro en la celda numero 3
        cont = 3
        i = 1
        
        for query in queryHistorial:

            fecha_pago = query['fechaPago']
            fecha_formateada = fecha_pago.strftime("%d/%m/%Y") if fecha_pago else ""
            #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws.cell(row = cont, column = 1).value = i                    
            ws.cell(row = cont, column = 2).value = int(query['asociado__numDocumento'])
            ws.cell(row = cont, column = 3).value = f"{query['asociado__nombre']} {query['asociado__apellido']}"
            ws.cell(row = cont, column = 4).value = query['formaPago_id__formaPago']
            ws.cell(row = cont, column = 5).value = query['total_pagado']
            ws.cell(row = cont, column = 6).value = fecha_formateada
        
            i+=1
            cont+=1

        nombre_archivo = f"Reporte_Conciliacion_Bancaria.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class DescargarAsociados(BaseReporteExcel):
    nombre_hoja = "Listado Asociados"
    columnas = [
        'ID Asociado', 'Nombres', 'Apellidos', 'Número Documento', 'Genero', 'Estado Civil', 'Tipo Vivienda', 'Estrato', 'Dirección', 'Barrio', 'Departamento Residencia', 'Municipio Residencia', 'Fecha Nacimiento', 'Indicativo Celular', 'Número Celular', 'Email', 'Estado Asociado', 'Tipo Asociado', 'Fecha Ingreso', 'Fecha Primer Cobro', 'Funeraria', 'Fecha Retiro'
        ]

    ancho_columnas = [11, 20, 20, 20, 20, 20, 20, 10, 25, 25, 14, 20, 15, 12, 15, 20, 18, 20, 20, 20, 22, 20]

    def get_queryset(self, request, *args, **kwargs):
        
        self.titulo = "Listado de Asociados"

        return (Asociado.objects.all()
                .select_related('deptoResidencia','mpioResidencia','tAsociado','indicativoCelular')
                .annotate(funeraria=F('parametroasociado__funeraria__concepto'), primerMes=F('parametroasociado__primerMes__concepto'))
                .order_by('pk'))

    def preparar_fila(self, obj):
        return [
            obj.pk,
            obj.nombre,
            obj.apellido,
            int(obj.numDocumento),
            obj.genero,
            obj.estadoCivil,
            obj.tipoVivienda,
            obj.estrato,
            obj.direccion,
            obj.barrio,
            obj.deptoResidencia.nombre,
            obj.mpioResidencia.nombre,
            obj.fechaNacimiento.strftime("%d/%m/%Y") if obj.fechaNacimiento else None,
            getattr(obj.indicativoCelular, 'indicativo', ''),
            int(obj.numCelular),
            obj.email,
            obj.estadoAsociado, 
            obj.tAsociado.concepto,
            obj.fechaIngreso.strftime("%d/%m/%Y") if obj.fechaIngreso else None,
            obj.primerMes,
            obj.funeraria,
            obj.fechaRetiro.strftime("%d/%m/%Y") if obj.fechaRetiro else None,
        ]


class DescargarTarifasAsociados(BaseReporteExcel):
    nombre_hoja = "Listado Tarifas Asociados"
    columnas = [
        'ID Asociado', 'Número Documento', 'Nombre Completo', 'Tipo Asociado', 'Estado Asociado', 'Valor', 'Aporte', 'Bienestar Social', 'Mascota', 'Repatriación', 'Seguro Vida', 'Adicionales', 'Coohoperativitos Aporte', 'Coohoperativitos B Social', 'Convenio'
        ]

    ancho_columnas = [11, 14, 36, 20, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]

    def get_queryset(self, request, *args, **kwargs):
        
        self.titulo = "Listado Tarifas de Asociados"

        return TarifaAsociado.objects.values('asociado__id',
                            'asociado__nombre','asociado__apellido','asociado__numDocumento','asociado__tAsociado__concepto','asociado__estadoAsociado','cuotaAporte','cuotaBSocial', 'cuotaMascota', 'cuotaRepatriacionBeneficiarios', 'cuotaRepatriacionTitular',
                            'cuotaSeguroVida', 'cuotaAdicionales', 'cuotaCoohopAporte', 'cuotaCoohopBsocial', 'cuotaConvenio', 'total'
                        )

    def preparar_fila(self, obj):
        return [
            obj['asociado__id'],
            int(obj['asociado__numDocumento']),
            f"{obj['asociado__nombre']} {obj['asociado__apellido']}",
            obj['asociado__tAsociado__concepto'],
            obj['asociado__estadoAsociado'],
            obj['total'],
            obj['cuotaAporte'],
            obj['cuotaBSocial'],
            obj['cuotaMascota'],
            (obj['cuotaRepatriacionBeneficiarios'] or 0) + (obj['cuotaRepatriacionTitular'] or 0),
            obj['cuotaSeguroVida'],
            obj['cuotaAdicionales'],
            obj['cuotaCoohopAporte'],
            obj['cuotaCoohopBsocial'],
            obj['cuotaConvenio'],
        ]


class DescargarBeneficiarios(BaseReporteExcel):
    nombre_hoja = "Listado Beneficiarios"
    columnas = [
        'ID Titular', 'Número Documento Titular', 'Nombre Titular', 'Apellido Titular', 'ID Beneficiario','Número Documento Beneficiario', 'Tipo Documento', 'Nombre Beneficiario', 'Apellido Beneficiario','Fecha Nacimiento', 'Parentesco', 'Repatriación'
        ]
    
    ancho_columnas = [11, 20, 24, 24, 13, 20, 20, 24, 24, 20, 20, 20]

    def get_queryset(self, request, *args, **kwargs):
        
        self.titulo = "Listado Beneficiarios"

        return Beneficiario.objects.filter(estadoRegistro=True).values(
                                        'asociado__id',
                                        'asociado__nombre',
                                        'asociado__apellido',
                                        'asociado__numDocumento',
                                        'id',
                                        'nombre',
                                        'apellido',
                                        'numDocumento',
                                        'tipoDocumento',
                                        'fechaNacimiento',
                                        'parentesco__nombre',
                                        'repatriacion',
                                        'paisRepatriacion__nombre',
                                        'ciudadRepatriacion'
                                    ).order_by('asociado')

    def preparar_fila(self, obj):
        repatriacion = ''
        if obj['repatriacion']:
            repatriacion = f"{obj['paisRepatriacion__nombre']} - {obj['ciudadRepatriacion']}"

        return [
            obj['asociado__id'],  # Código titular
            int(obj['asociado__numDocumento']),  # Documento titular
            obj['asociado__nombre'],  # Nombres titular
            obj['asociado__apellido'],  # Apellidos titular
            obj['id'],
            int(obj['numDocumento']),  # Documento beneficiario
            obj['tipoDocumento'],
            obj['nombre'],
            obj['apellido'], 
            obj['fechaNacimiento'].strftime("%d/%m/%Y"),
            obj['parentesco__nombre'],
            repatriacion
        ]


class DescargarMascotas(BaseReporteExcel):
    nombre_hoja = "Listado Mascotas"
    columnas = [
        'ID Titular', 'Número Documento Titular', 'Nombre Titular', 'Apellido Titular',  'ID Mascota', 'Nombre Mascota', 'Tipo', 'Raza', 'Fecha Nacimiento'
        ]
    
    ancho_columnas = [11, 20, 24, 24, 11, 24, 20, 20, 24]

    def get_queryset(self, request, *args, **kwargs):
        
        self.titulo = "Listado Mascotas"

        return Mascota.objects.filter(estadoRegistro = True).values(
                                            'asociado__id',
                                            'asociado__nombre',
                                            'asociado__apellido',
                                            'asociado__numDocumento',
                                            'nombre',
                                            'id',
                                            'tipo',
                                            'raza',
                                            'fechaNacimiento',
                                        ).order_by('asociado')

    def preparar_fila(self, obj):

        return [
            obj['asociado__id'],
            int(obj['asociado__numDocumento']),
            obj['asociado__nombre'],
            obj['asociado__apellido'],
            obj['id'],
            obj['nombre'],
            obj['tipo'],
            obj['raza'], 
            obj['fechaNacimiento'].strftime("%d/%m/%Y"),
        ]

class DescargarAdicionalesFuneraria(BaseReporteExcel):
    nombre_hoja = "Listado Adicionales Funeraria"
    columnas = [
        'ID Titular', 'Número Documento Titular', 'Nombre Titular', 'Apellido Titular',  'Concepto', 'Fecha Inicio', 'Cuota Adicionales'
        ]
    
    ancho_columnas = [11, 20, 24, 24, 34, 20, 20]

    def get_queryset(self, request, *args, **kwargs):
        
        self.titulo = "Listado Adicionales Funeraria"

        return TarifaAsociado.objects.filter(estadoRegistro = True, estadoAdicional = True).values(
                                            'asociado__id',
                                            'asociado__nombre',
                                            'asociado__apellido',
                                            'asociado__numDocumento',
                                            'conceptoAdicional',
                                            'fechaInicioAdicional',
                                            'cuotaAdicionales'
                                        ).order_by('asociado')

    def preparar_fila(self, obj):

        return [
            obj['asociado__id'],
            int(obj['asociado__numDocumento']),
            obj['asociado__nombre'],
            obj['asociado__apellido'],
            obj['conceptoAdicional'],
            obj['fechaInicioAdicional'].strftime("%d/%m/%Y"),
            int(obj['cuotaAdicionales']), 
        ]

class DescargarAuxilios(BaseReporteExcel):
    nombre_hoja = "Listado Auxilios"
    columnas = [
        'ID Auxilio', 'Número Documento', 'Nombre Completo', 'Tipo Auxilio', 'Entidad Bancaria', 'Número Cuenta', 'Valor Auxilio', 'Fecha Solicitud', 'Estado', 'Observaciones', 'Fecha Desembolso'
        ]
    
    ancho_columnas = [14, 20, 37, 32, 25, 17, 13, 13, 20, 40, 20]

    def get(self, request, *args, **kwargs):
        template = 'reporte/modalReporte.html'
        return render(request, template, {'tipoReporte':'auxilio'})

    def post(self, request, *args, **kwargs):
        return self.exportar_excel(request, *args, **kwargs)

    def get_queryset(self, request, *args, **kwargs):
        fechaInicialForm = request.POST.get('fechaInicial')
        fechaFinalForm = request.POST.get('fechaFinal')

        if not fechaInicialForm or not fechaFinalForm:
            raise ValueError("Fechas no enviadas en el formulario.")
    
        fechaInicial = datetime.strptime(fechaInicialForm, "%Y-%m-%d")
        fechaFinal = datetime.strptime(fechaFinalForm, "%Y-%m-%d")

        self.titulo = f"Reporte de Auxilios desde {fechaInicial.strftime('%d-%m-%Y')} hasta {fechaFinal.strftime('%d-%m-%Y')}"

        return HistoricoAuxilio.objects.filter(
            fechaSolicitud__range=[fechaInicialForm, fechaFinalForm]
        ).select_related('asociado','tipoAuxilio')

    def preparar_fila(self, obj):
        return [
            obj.id,
            int(obj.asociado.numDocumento),
            f'{obj.asociado.nombre} {obj.asociado.apellido}',
            obj.tipoAuxilio.nombre,
            obj.entidadBancaria,
            obj.numCuenta,
            obj.valor,
            obj.fechaSolicitud.strftime("%d/%m/%Y"),
            obj.estado,
            obj.observacion,
            obj.fechaDesembolso.strftime("%d/%m/%Y") if obj.fechaDesembolso else '',
        ]

class DescargarCreditos(BaseReporteExcel):
    nombre_hoja = "Listado Créditos"
    columnas = [
        'ID Crédito', 'Número Documento', 'Nombre Completo', 'Fecha Solicitud', 'Valor', 'Número Cuotas', 'Valor Cuota', 'Total Credito', 'Cuotas Pagas', 'Pendiente Pago', 'Estado', 'Linea Credito', 'Amortización', 'Tasa Interes', 'Forma Desembolso', 'Medio de Pago', 'Tipo Asociado', 'Banco', 'Tipo de Cuenta', 'Número de Cuenta'
        ]
    
    ancho_columnas = [11, 14, 32, 15, 12, 12, 12, 13, 13, 13, 13, 21, 16, 20, 28, 15, 21, 16, 18, 15]

    def get(self, request, *args, **kwargs):
        template = 'reporte/modalReporte.html'
        return render(request, template, {'tipoReporte':'credito'})

    def post(self, request, *args, **kwargs):
        return self.exportar_excel(request, *args, **kwargs)

    def get_queryset(self, request, *args, **kwargs):
        fechaInicialForm = request.POST.get('fechaInicial')
        fechaFinalForm = request.POST.get('fechaFinal')

        if not fechaInicialForm or not fechaFinalForm:
            raise ValueError("Fechas no enviadas en el formulario.")
    
        fechaInicial = datetime.strptime(fechaInicialForm, "%Y-%m-%d")
        fechaFinal = datetime.strptime(fechaFinalForm, "%Y-%m-%d")

        self.titulo = f"Reporte de Créditos desde {fechaInicial.strftime('%d-%m-%Y')} hasta {fechaFinal.strftime('%d-%m-%Y')}"

        return HistoricoCredito.objects.filter(
            fechaSolicitud__range=[fechaInicialForm, fechaFinalForm]
        ).select_related('asociado__tAsociado','tasaInteres')

    def preparar_fila(self, obj):
        return [
            obj.id,
            int(obj.asociado.numDocumento),
            f'{obj.asociado.nombre} {obj.asociado.apellido}',
            obj.fechaSolicitud.strftime("%d/%m/%Y"),
            obj.valor,
            obj.cuotas,
            obj.valorCuota,
            obj.totalCredito,
            obj.cuotasPagas,
            obj.pendientePago,
            obj.estado,
            obj.lineaCredito,
            obj.amortizacion,
            obj.tasaInteres.concepto if obj.tasaInteres else 'Sin Tasa',
            obj.formaDesembolso,
            obj.medioPago,
            obj.asociado.tAsociado.concepto,
            obj.banco,
            obj.tipoCuenta,
            obj.numCuenta,
        ]

class DescargarVentasHE(BaseReporteExcel):

    nombre_hoja = "Ventas Home Elements"
    columnas = [
        'ID Venta', 'Fecha Venta', 'Número Documento', 'Nombre Completo', 'Forma de Pago', 'Cuotas', 'Valor Cuotas', 'Cuotas Pagadas', 'Pendiente Pago', 'Valor Neto', 'Tasa Interes', 'Productos', 'Cantidad', 'Precio', 'Total Bruto', 'Total Neto'
        ]
    ancho_columnas = [11, 14, 14 , 32, 20, 14, 12, 12, 12, 12, 25, 20, 12, 13, 13, 13]

    def get(self, request, *args, **kwargs):
        template = 'reporte/modalReporte.html'
        return render(request, template, {'tipoReporte':'ventasHE'})
    
    def post(self, request, *args, **kwargs):
        return self.exportar_excel(request, *args, **kwargs)

    def get_queryset(self, request, *args, **kwargs):
        fechaInicialForm = request.POST.get('fechaInicial')
        fechaFinalForm = request.POST.get('fechaFinal')

        if not fechaInicialForm or not fechaFinalForm:
            raise ValueError("Fechas no enviadas en el formulario.")
    
        fechaInicial = datetime.strptime(fechaInicialForm, "%Y-%m-%d")
        fechaFinal = datetime.strptime(fechaFinalForm, "%Y-%m-%d")

        self.titulo = f"Reporte de Pagos desde {fechaInicial.strftime('%d-%m-%Y')} hasta {fechaFinal.strftime('%d-%m-%Y')}"

        return HistoricoVenta.objects.filter(
                    fechaVenta__range=[fechaInicial, fechaFinal],
                    estadoRegistro=True
                ).select_related('asociado','tasaInteres'
                ).prefetch_related(
                    Prefetch('detalleventa_set', queryset=DetalleVenta.objects.select_related('producto'))
                ).order_by('pk')

    def preparar_fila(self, obj):
        return [
            obj.id,
            obj.fechaVenta.strftime("%d/%m/%Y"),
            int(obj.asociado.numDocumento),
            f'{obj.asociado.nombre} {obj.asociado.apellido}',
            obj.formaPago,
            obj.cuotas,
            obj.valorCuotas,
            obj.cuotasPagas,
            obj.pendientePago,
            obj.valorNeto,
            obj.tasaInteres.concepto if obj.tasaInteres else 'Sin Tasa',
        ]
    
    def generar_excel(self, queryset, titulo):
        wb = Workbook()
        ws = wb.active
        ws.title = self.nombre_hoja

        # Estilos
        bold_font = Font(bold=True, size=16, color="FFFFFF")
        bold_font_header = Font(bold=True, size=11, color="000000")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_header = PatternFill(start_color="85B84C", end_color="85B84C", fill_type="solid")

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.columnas))
        celda_titulo = ws.cell(row=1, column=1, value=titulo)
        celda_titulo.font = bold_font
        celda_titulo.alignment = alignment_center
        celda_titulo.fill = fill_header

        # Encabezados
        for idx, encabezado in enumerate(self.columnas, 1):
            celda = ws.cell(row=2, column=idx)
            celda.value = encabezado
            celda.font = bold_font_header
            celda.alignment = alignment_center

        # Ancho de columnas
        for idx, width in enumerate(self.ancho_columnas, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

        # Filas de contenido
        fila_actual = 3
        for venta in queryset:
            datos_venta = self.preparar_fila(venta)
            productos = venta.detalleventa_set.all()

            for idx, producto in enumerate(productos):
                datos_producto = [
                    producto.producto.nombre,
                    producto.cantidad,
                    producto.precio,
                    producto.totalBruto,
                    producto.totalNeto
                ]
                # Repetir solo el ID de la venta, el resto vacío si no es la primera fila
                datos_venta_parciales = (
                    [datos_venta[0]] + [''] * (len(datos_venta) - 1) if idx > 0 else datos_venta
                )
                fila_completa = datos_venta_parciales + datos_producto

                for col, valor in enumerate(fila_completa, 1):
                    ws.cell(row=fila_actual, column=col).value = valor

                fila_actual += 1

            # Si no hay productos, mostrar solo datos de venta
            if not productos:
                for col, valor in enumerate(datos_venta, 1):
                    ws.cell(row=fila_actual, column=col).value = valor
                fila_actual += 1

        return wb   

class DescargarProductosHE(BaseReporteExcel):
    nombre_hoja = "Productos Home Elements"
    columnas = [
        'ID Producto', 'Rerencia', 'Código EAN', 'Nombre', 'Precio', 'Categoria', 'Proveedor', 'Descripcion'
        ]
    ancho_columnas = [11, 14, 14, 32, 14, 22, 17, 60]

    def get_queryset(self, request, *args, **kwargs):
        
        self.titulo = "Listado de Productos Home Elements"

        return Producto.objects.all().select_related('categoria','proveedor').order_by('pk')

    def preparar_fila(self, obj):
        return [
            obj.id,
            obj.referencia,
            obj.ean,
            obj.nombre,
            obj.precio,
            obj.categoria.nombre,
            obj.proveedor.razonSocial,
            obj.descripcion,
        ]